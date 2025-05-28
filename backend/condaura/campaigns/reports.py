import io
import csv
import datetime
import pandas as pd
from django.http import HttpResponse
from django.template.loader import get_template
from xhtml2pdf import pisa
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from .models import Campaign
from access.models import Review

class ReportGenerator:
    @staticmethod
    def generate_excel_report(campaign_id):
        """
        Génère un rapport Excel pour une campagne
        """
        try:
            campaign = Campaign.objects.get(id=campaign_id)
            
            # Créer un workbook et une feuille de calcul
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Campagne"
            
            # En-tête avec informations de la campagne
            worksheet.merge_cells('A1:F1')
            worksheet['A1'] = f"Rapport de campagne: {campaign.name}"
            worksheet['A1'].font = Font(bold=True, size=14)
            worksheet['A1'].alignment = Alignment(horizontal='center')
            
            worksheet.merge_cells('A2:F2')
            worksheet['A2'] = f"Période: {campaign.start_date.strftime('%d/%m/%Y')} - {campaign.end_date.strftime('%d/%m/%Y')}"
            worksheet['A2'].alignment = Alignment(horizontal='center')
            
            worksheet.merge_cells('A3:F3')
            worksheet['A3'] = f"Statut: {campaign.status} - Progression: {campaign.progress}%"
            worksheet['A3'].alignment = Alignment(horizontal='center')
            
            # En-tête du tableau des revues
            headers = ["Utilisateur", "Département", "Ressource", "Type", "Niveau d'accès", "Décision", "Commentaire", "Réviseur", "Date de revue"]
            for col_num, header in enumerate(headers, 1):
                cell = worksheet.cell(row=5, column=col_num)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
            
            # Contenu du tableau
            reviews = Review.objects.filter(campaign=campaign).select_related('access', 'access__user', 'reviewer')
            
            for row_num, review in enumerate(reviews, 6):
                worksheet.cell(row=row_num, column=1).value = f"{review.access.user.first_name} {review.access.user.last_name}"
                worksheet.cell(row=row_num, column=2).value = review.access.user.department
                worksheet.cell(row=row_num, column=3).value = review.access.resource_name
                worksheet.cell(row=row_num, column=4).value = review.access.resource_type
                worksheet.cell(row=row_num, column=5).value = review.access.access_level
                worksheet.cell(row=row_num, column=6).value = review.get_decision_display()
                worksheet.cell(row=row_num, column=7).value = review.comment
                worksheet.cell(row=row_num, column=8).value = f"{review.reviewer.first_name} {review.reviewer.last_name}"
                worksheet.cell(row=row_num, column=9).value = review.reviewed_at.strftime('%d/%m/%Y %H:%M') if review.reviewed_at else "Non revu"
            
            # Ajuster la largeur des colonnes
            for col_num, _ in enumerate(headers, 1):
                worksheet.column_dimensions[get_column_letter(col_num)].width = 20
            
            # Statistiques
            row_num = len(reviews) + 8
            
            worksheet.merge_cells(f'A{row_num}:F{row_num}')
            worksheet[f'A{row_num}'] = "Statistiques"
            worksheet[f'A{row_num}'].font = Font(bold=True, size=12)
            worksheet[f'A{row_num}'].alignment = Alignment(horizontal='center')
            
            # Statistiques par décision
            decision_stats = {}
            for review in reviews:
                decision = review.get_decision_display()
                if decision not in decision_stats:
                    decision_stats[decision] = 0
                decision_stats[decision] += 1
            
            row_num += 2
            worksheet[f'A{row_num}'] = "Décisions"
            worksheet[f'A{row_num}'].font = Font(bold=True)
            
            for decision, count in decision_stats.items():
                row_num += 1
                worksheet[f'A{row_num}'] = decision
                worksheet[f'B{row_num}'] = count
                worksheet[f'C{row_num}'] = f"{count / len(reviews) * 100:.1f}%"
            
            # Statistiques par type de ressource
            resource_stats = {}
            for review in reviews:
                resource_type = review.access.resource_type
                if resource_type not in resource_stats:
                    resource_stats[resource_type] = 0
                resource_stats[resource_type] += 1
            
            row_num += 2
            worksheet[f'A{row_num}'] = "Types de ressources"
            worksheet[f'A{row_num}'].font = Font(bold=True)
            
            for resource_type, count in resource_stats.items():
                row_num += 1
                worksheet[f'A{row_num}'] = resource_type
                worksheet[f'B{row_num}'] = count
                worksheet[f'C{row_num}'] = f"{count / len(reviews) * 100:.1f}%"
            
            # Créer le fichier en mémoire
            output = io.BytesIO()
            workbook.save(output)
            output.seek(0)
            
            # Générer la réponse HTTP
            response = HttpResponse(
                output.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename=campagne_{campaign.id}_{datetime.date.today().isoformat()}.xlsx'
            
            return response
        
        except Campaign.DoesNotExist:
            return None
        except Exception as e:
            print(f"Error generating Excel report: {str(e)}")
            return None
    
    @staticmethod
    def generate_pdf_report(campaign_id):
        """
        Génère un rapport PDF pour une campagne
        """
        try:
            campaign = Campaign.objects.get(id=campaign_id)
            
            # Récupérer les revues
            reviews = Review.objects.filter(campaign=campaign).select_related('access', 'access__user', 'reviewer')
            
            # Calcul des statistiques
            total_reviews = len(reviews)
            approved = sum(1 for r in reviews if r.decision == 'approved')
            rejected = sum(1 for r in reviews if r.decision == 'rejected')
            pending = sum(1 for r in reviews if r.decision == 'pending')
            
            # Statistiques par type de ressource
            resource_stats = {}
            for review in reviews:
                resource_type = review.access.resource_type
                if resource_type not in resource_stats:
                    resource_stats[resource_type] = {
                        'total': 0,
                        'approved': 0,
                        'rejected': 0,
                        'pending': 0
                    }
                resource_stats[resource_type]['total'] += 1
                resource_stats[resource_type][review.decision] += 1
            
            # Préparer le contexte pour le template
            context = {
                'campaign': campaign,
                'reviews': reviews,
                'total_reviews': total_reviews,
                'approved': approved,
                'rejected': rejected,
                'pending': pending,
                'approved_percent': round(approved / total_reviews * 100, 1) if total_reviews else 0,
                'rejected_percent': round(rejected / total_reviews * 100, 1) if total_reviews else 0,
                'pending_percent': round(pending / total_reviews * 100, 1) if total_reviews else 0,
                'resource_stats': resource_stats,
                'date_generated': datetime.datetime.now().strftime('%d/%m/%Y %H:%M')
            }
            
            # Rendre le template HTML
            template = get_template('campaign_report.html')
            html = template.render(context)
            
            # Créer le PDF
            result = io.BytesIO()
            pdf = pisa.pisaDocument(io.BytesIO(html.encode("UTF-8")), result)
            
            if not pdf.err:
                # Générer la réponse HTTP
                response = HttpResponse(result.getvalue(), content_type='application/pdf')
                response['Content-Disposition'] = f'attachment; filename=campagne_{campaign.id}_{datetime.date.today().isoformat()}.pdf'
                return response
            
            return None
        
        except Campaign.DoesNotExist:
            return None
        except Exception as e:
            print(f"Error generating PDF report: {str(e)}")
            return None
    
    @staticmethod
    def generate_csv_report(campaign_id):
        """
        Génère un rapport CSV pour une campagne
        """
        try:
            campaign = Campaign.objects.get(id=campaign_id)
            
            # Récupérer les revues
            reviews = Review.objects.filter(campaign=campaign).select_related('access', 'access__user', 'reviewer')
            
            # Créer le fichier CSV en mémoire
            output = io.StringIO()
            writer = csv.writer(output)
            
            # En-tête
            writer.writerow([
                'Utilisateur', 'Email', 'Département', 'Ressource', 'Type', 'Niveau d\'accès',
                'Décision', 'Commentaire', 'Réviseur', 'Date de revue'
            ])
            
            # Contenu
            for review in reviews:
                writer.writerow([
                    f"{review.access.user.first_name} {review.access.user.last_name}",
                    review.access.user.email,
                    review.access.user.department,
                    review.access.resource_name,
                    review.access.resource_type,
                    review.access.access_level,
                    review.get_decision_display(),
                    review.comment,
                    f"{review.reviewer.first_name} {review.reviewer.last_name}",
                    review.reviewed_at.strftime('%d/%m/%Y %H:%M') if review.reviewed_at else "Non revu"
                ])
            
            # Générer la réponse HTTP
            response = HttpResponse(output.getvalue(), content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename=campagne_{campaign.id}_{datetime.date.today().isoformat()}.csv'
            
            return response
        
        except Campaign.DoesNotExist:
            return None
        except Exception as e:
            print(f"Error generating CSV report: {str(e)}")
            return None